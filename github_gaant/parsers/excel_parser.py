"""Excel parser for Gaant files."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import IssueState, Project, Task


# Column definitions
COLUMNS = [
    ("issue", "Issue #", 10),
    ("title", "Title", 40),
    ("start", "Start Date", 15),
    ("end", "End Date", 15),
    ("assignees", "Assignees", 20),
    ("labels", "Labels", 20),
    ("status", "Status", 12),
    ("progress", "Progress %", 12),
    ("parent", "Parent #", 10),
    ("milestone", "Milestone", 15),
]


def _parse_date(value: Any) -> Optional[date]:
    """Parse a date value from Excel."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _parse_list(value: Any) -> List[str]:
    """Parse a comma-separated list from Excel."""
    if value is None:
        return []
    if isinstance(value, str):
        return [v.strip() for v in value.split(",") if v.strip()]
    return []


def _format_list(items: List[str]) -> str:
    """Format a list as comma-separated string."""
    return ", ".join(items)


def task_to_row(task: Task, parent_issue: Optional[int] = None) -> Dict[str, Any]:
    """Convert a Task to a dictionary for Excel row."""
    return {
        "issue": task.issue_number,
        "title": task.title,
        "start": task.start_date,
        "end": task.end_date,
        "assignees": _format_list(task.assignees),
        "labels": _format_list(task.labels),
        "status": "closed" if task.state == IssueState.CLOSED else "open",
        "progress": task.progress,
        "parent": parent_issue,
        "milestone": task.milestone,
    }


def flatten_tasks_for_excel(tasks: List[Task], parent_issue: Optional[int] = None) -> List[Dict[str, Any]]:
    """Flatten task hierarchy into rows for Excel."""
    rows: List[Dict[str, Any]] = []
    for task in tasks:
        rows.append(task_to_row(task, parent_issue))
        if task.subtasks:
            rows.extend(flatten_tasks_for_excel(task.subtasks, task.issue_number))
    return rows


def save_project_to_excel(project: Project, path: Path) -> None:
    """Save a Project to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Write headers
    for col_idx, (key, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Flatten tasks and write rows
    rows = flatten_tasks_for_excel(project.tasks)
    
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            value = row_data.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format dates
            if key in ("start", "end") and isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
            
            # Indent subtasks
            if key == "title" and row_data.get("parent"):
                cell.alignment = Alignment(indent=2)
    
    # Add project info sheet
    ws_info = wb.create_sheet("Project Info")
    info_data = [
        ("Project ID", project.id),
        ("Project Number", project.number),
        ("Project Title", project.title),
        ("Project URL", project.url),
        ("Overall Progress", f"{project.overall_progress}%"),
        ("Total Tasks", project.total_tasks),
    ]
    for row_idx, (label, value) in enumerate(info_data, start=1):
        ws_info.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_info.cell(row=row_idx, column=2, value=value)
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 50
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(path)


def row_to_task(row_data: Dict[str, Any]) -> Task:
    """Convert an Excel row to a Task."""
    return Task(
        issue_number=int(row_data.get("issue") or 0),
        issue_id=str(row_data.get("issue_id", "")),
        project_item_id=row_data.get("project_item_id"),
        title=str(row_data.get("title", "")),
        body=row_data.get("body"),
        state=IssueState.CLOSED if row_data.get("status") == "closed" else IssueState.OPEN,
        start_date=_parse_date(row_data.get("start")),
        end_date=_parse_date(row_data.get("end")),
        assignees=_parse_list(row_data.get("assignees")),
        labels=_parse_list(row_data.get("labels")),
        milestone=row_data.get("milestone") if row_data.get("milestone") else None,
        parent_issue_number=int(row_data.get("parent")) if row_data.get("parent") else None,
    )


def load_project_from_excel(path: Path) -> Project:
    """Load a Project from an Excel file."""
    wb = load_workbook(path)
    ws = wb["Tasks"]
    
    # Read headers
    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx for idx, header in enumerate(headers)}
    
    # Read rows
    all_tasks: List[Task] = []
    task_by_issue: Dict[int, Task] = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
        
        row_data = {}
        for key, header, _ in COLUMNS:
            if header in col_map:
                row_data[key] = row[col_map[header]]
        
        task = row_to_task(row_data)
        all_tasks.append(task)
        if task.issue_number > 0:
            task_by_issue[task.issue_number] = task
    
    # Build hierarchy
    root_tasks: List[Task] = []
    for task in all_tasks:
        if task.parent_issue_number and task.parent_issue_number in task_by_issue:
            parent = task_by_issue[task.parent_issue_number]
            parent.subtasks.append(task)
        else:
            root_tasks.append(task)
    
    # Try to read project info
    project_id = ""
    project_number = 0
    project_title = ""
    project_url = None
    
    if "Project Info" in wb.sheetnames:
        ws_info = wb["Project Info"]
        info_dict = {}
        for row in ws_info.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0]:
                info_dict[row[0]] = row[1]
        
        project_id = str(info_dict.get("Project ID", ""))
        project_number = int(info_dict.get("Project Number", 0) or 0)
        project_title = str(info_dict.get("Project Title", ""))
        project_url = info_dict.get("Project URL")
    
    return Project(
        id=project_id,
        number=project_number,
        title=project_title,
        url=project_url,
        tasks=root_tasks,
    )


def excel_exists(path: Path) -> bool:
    """Check if the Excel file exists."""
    return path.exists() and path.is_file() and path.suffix in (".xlsx", ".xls")
